#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
بوت تيليجرام لإدارة تقارير زيارات المدارس
School Inspection Telegram Bot - Fixed Version
"""

import os
import logging
import sqlite3
from datetime import datetime, timedelta
from typing import Dict, Any

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# إعداد السجلات
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# قراءة الإعدادات من Environment Variables مباشرة
BOT_TOKEN = os.environ.get('BOT_TOKEN')
GROUP_CHAT_ID = os.environ.get('GROUP_CHAT_ID')
ADMIN_IDS = [int(id.strip()) for id in os.environ.get('ADMIN_IDS', '').split(',') if id.strip()]
SUPERVISOR_IDS = [int(id.strip()) for id in os.environ.get('SUPERVISOR_IDS', '').split(',') if id.strip()]

# أسماء المشرفين
SUPERVISORS_NAMES = [
    "ممدوح", "افنان", "عبدالله", "ريان", 
    "مصطفى", "موسى", "طه", "محمد"
]

# حالات المحادثة
(
    SUPERVISOR_NAME, VISIT_DATE, SCHOOL_NAME,
    MAINTENANCE_NOTES, AC_NOTES, CLEANING_NOTES,
    REVIEW_REPORT
) = range(7)

# قاعدة البيانات
DB_NAME = 'school_reports.db'


def init_database():
    """إنشاء قاعدة البيانات والجداول"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            supervisor_name TEXT NOT NULL,
            visit_date DATE NOT NULL,
            school_name TEXT NOT NULL,
            maintenance_notes TEXT,
            ac_notes TEXT,
            cleaning_notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()
    logger.info("تم إنشاء قاعدة البيانات بنجاح")


def save_report(user_id: int, data: Dict[str, Any]) -> int:
    """حفظ التقرير في قاعدة البيانات"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO reports (
            user_id, supervisor_name, visit_date, school_name,
            maintenance_notes, ac_notes, cleaning_notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        user_id,
        data['supervisor_name'],
        data['visit_date'],
        data['school_name'],
        data['maintenance_notes'],
        data['ac_notes'],
        data['cleaning_notes']
    ))
    
    report_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return report_id


def get_reports_by_period(period: str) -> list:
    """استخراج التقارير حسب الفترة المحددة"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    today = datetime.now().date()
    
    if period == 'today':
        start_date = today
        end_date = today
    elif period == 'week':
        # الأسبوع يبدأ الجمعة وينتهي الخميس
        days_since_friday = (today.weekday() + 3) % 7
        start_date = today - timedelta(days=days_since_friday)
        end_date = start_date + timedelta(days=6)
    elif period == 'month':
        start_date = today.replace(day=1)
        if today.month == 12:
            end_date = today.replace(day=31)
        else:
            end_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
    else:
        return []
    
    cursor.execute('''
        SELECT supervisor_name, visit_date, school_name,
               maintenance_notes, ac_notes, cleaning_notes
        FROM reports
        WHERE visit_date BETWEEN ? AND ?
        ORDER BY visit_date, supervisor_name
    ''', (start_date, end_date))
    
    reports = cursor.fetchall()
    conn.close()
    
    return reports


def create_excel_report(period: str, reports: list) -> str:
    """إنشاء ملف Excel للتقارير"""
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    sections = [
        ('الصيانة', 3),
        ('التكييف', 4),
        ('النظافة', 5)
    ]
    
    for section_name, notes_column in sections:
        ws = wb.create_sheet(title=section_name)
        
        headers = ['التاريخ', 'المشرف', 'المدرسة', 'الملاحظة']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for report in reports:
            visit_date = report[1]
            supervisor_name = report[0]
            school_name = report[2]
            note = report[notes_column]
            
            if not note or note.strip() == '':
                note = 'لا يوجد'
            
            ws.append([visit_date, supervisor_name, school_name, note])
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    
    summary_sheet = wb.create_sheet(title='الملخص', index=0)
    summary_sheet.append(['نوع التقرير', 'الفترة'])
    summary_sheet.append(['إجمالي التقارير', len(reports)])
    summary_sheet.append([''])
    summary_sheet.append(['القسم', 'عدد الملاحظات'])
    
    for section_name, notes_column in sections:
        count = sum(1 for r in reports if r[notes_column] and r[notes_column].strip() and r[notes_column].strip().lower() != 'لا يوجد')
        summary_sheet.append([section_name, count])
    
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 20
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'تقرير_{period}_{timestamp}.xlsx'
    wb.save(filename)
    
    return filename


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """أمر البداية"""
    user_id = update.effective_user.id
    
    welcome_text = "🏫 مرحباً بك في بوت تقارير المدارس\n\n"
    
    if user_id in SUPERVISOR_IDS:
        welcome_text += "أنت مشرف ميداني ✅\n\n"
        welcome_text += "الأوامر المتاحة:\n"
        welcome_text += "/report - إرسال تقرير زيارة جديد\n"
        welcome_text += "/cancel - إلغاء العملية الحالية"
    elif user_id in ADMIN_IDS:
        welcome_text += "أنت مدير النظام 👨‍💼\n\n"
        welcome_text += "الأوامر المتاحة:\n"
        welcome_text += "/summary - استخراج تقرير Excel\n"
        welcome_text += "/summary_today - تقرير اليوم\n"
        welcome_text += "/summary_week - تقرير الأسبوع\n"
        welcome_text += "/summary_month - تقرير الشهر"
    else:
        welcome_text += "⚠️ عذراً، ليس لديك صلاحية استخدام هذا البوت"
    
    await update.message.reply_text(welcome_text)


async def start_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """بدء إدخال تقرير جديد"""
    user_id = update.effective_user.id
    
    if user_id not in SUPERVISOR_IDS:
        await update.message.reply_text("⚠️ عذراً، هذا الأمر متاح للمشرفين الميدانيين فقط")
        return ConversationHandler.END
    
    keyboard = []
    for i in range(0, len(SUPERVISORS_NAMES), 2):
        row = [KeyboardButton(SUPERVISORS_NAMES[i])]
        if i + 1 < len(SUPERVISORS_NAMES):
            row.append(KeyboardButton(SUPERVISORS_NAMES[i + 1]))
        keyboard.append(row)
    
    keyboard.append([KeyboardButton("✍️ كتابة يدوياً")])
    
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text(
        "📝 إرسال تقرير زيارة جديد\n\n"
        "الخطوة 1️⃣: اختر اسم المشرف",
        reply_markup=reply_markup
    )
    
    return SUPERVISOR_NAME


async def get_supervisor_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """الحصول على اسم المشرف"""
    supervisor_name = update.message.text
    
    if supervisor_name == "✍️ كتابة يدوياً":
        await update.message.reply_text(
            "اكتب اسم المشرف:",
            reply_markup=ReplyKeyboardMarkup([[KeyboardButton("إلغاء")]], one_time_keyboard=True, resize_keyboard=True)
        )
        return SUPERVISOR_NAME
    
    if supervisor_name == "إلغاء":
        await update.message.reply_text("تم الإلغاء", reply_markup=ReplyKeyboardMarkup([[]], resize_keyboard=True))
        return ConversationHandler.END
    
    context.user_data['supervisor_name'] = supervisor_name
    
    keyboard = [
        [KeyboardButton("📅 اليوم"), KeyboardButton("📅 أمس")],
        [KeyboardButton("✍️ إدخال تاريخ")]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text(
        "الخطوة 2️⃣: اختر تاريخ الزيارة",
        reply_markup=reply_markup
    )
    
    return VISIT_DATE


async def get_visit_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """الحصول على تاريخ الزيارة"""
    date_input = update.message.text
    
    if date_input == "📅 اليوم":
        visit_date = datetime.now().date()
    elif date_input == "📅 أمس":
        visit_date = (datetime.now() - timedelta(days=1)).date()
    elif date_input == "✍️ إدخال تاريخ":
        await update.message.reply_text(
            "أدخل التاريخ بالصيغة: YYYY-MM-DD\n"
            "مثال: 2024-12-17",
            reply_markup=ReplyKeyboardMarkup([[]], resize_keyboard=True)
        )
        return VISIT_DATE
    else:
        try:
            visit_date = datetime.strptime(date_input, '%Y-%m-%d').date()
        except ValueError:
            await update.message.reply_text(
                "⚠️ صيغة تاريخ خاطئة. يرجى إدخال التاريخ بالصيغة: YYYY-MM-DD\n"
                "مثال: 2024-12-17"
            )
            return VISIT_DATE
    
    context.user_data['visit_date'] = str(visit_date)
    
    await update.message.reply_text(
        f"تاريخ الزيارة: {visit_date}\n\n"
        "الخطوة 3️⃣: أدخل اسم المدرسة",
        reply_markup=ReplyKeyboardMarkup([[]], resize_keyboard=True)
    )
    
    return SCHOOL_NAME


async def get_school_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """الحصول على اسم المدرسة"""
    context.user_data['school_name'] = update.message.text
    
    await update.message.reply_text(
        "الخطوة 4️⃣: أدخل ملاحظات الصيانة\n"
        "(إذا لم يكن هناك ملاحظات، اكتب: لا يوجد)"
    )
    
    return MAINTENANCE_NOTES


async def get_maintenance_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """الحصول على ملاحظات الصيانة"""
    context.user_data['maintenance_notes'] = update.message.text
    
    await update.message.reply_text(
        "الخطوة 5️⃣: أدخل ملاحظات التكييف\n"
        "(إذا لم يكن هناك ملاحظات، اكتب: لا يوجد)"
    )
    
    return AC_NOTES


async def get_ac_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """الحصول على ملاحظات التكييف"""
    context.user_data['ac_notes'] = update.message.text
    
    await update.message.reply_text(
        "الخطوة 6️⃣: أدخل ملاحظات النظافة\n"
        "(إذا لم يكن هناك ملاحظات، اكتب: لا يوجد)"
    )
    
    return CLEANING_NOTES


async def get_cleaning_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """الحصول على ملاحظات النظافة"""
    context.user_data['cleaning_notes'] = update.message.text
    
    data = context.user_data
    review_text = (
        "📋 مراجعة التقرير:\n\n"
        f"👤 المشرف: {data['supervisor_name']}\n"
        f"📅 التاريخ: {data['visit_date']}\n"
        f"🏫 المدرسة: {data['school_name']}\n\n"
        f"🔧 الصيانة:\n{data['maintenance_notes']}\n\n"
        f"❄️ التكييف:\n{data['ac_notes']}\n\n"
        f"🧹 النظافة:\n{data['cleaning_notes']}\n\n"
        "هل تريد اعتماد وإرسال التقرير؟"
    )
    
    keyboard = [
        [
            InlineKeyboardButton("✅ اعتماد وإرسال", callback_data='confirm_report'),
            InlineKeyboardButton("❌ إلغاء", callback_data='cancel_report')
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(review_text, reply_markup=reply_markup)
    
    return REVIEW_REPORT


async def confirm_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تأكيد وإرسال التقرير"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'cancel_report':
        await query.edit_message_text("❌ تم إلغاء التقرير")
        context.user_data.clear()
        return ConversationHandler.END
    
    user_id = update.effective_user.id
    report_id = save_report(user_id, context.user_data)
    
    await query.edit_message_text("✅ تم اعتماد التقرير بنجاح!")
    
    data = context.user_data
    group_message = (
        f"📊 تقرير زيارة — {data['visit_date']}\n"
        f"👤 المشرف: {data['supervisor_name']}\n"
        f"🏫 المدرسة: {data['school_name']}\n\n"
        f"🔧 الصيانة:\n{data['maintenance_notes']}\n\n"
        f"❄️ التكييف:\n{data['ac_notes']}\n\n"
        f"🧹 النظافة:\n{data['cleaning_notes']}"
    )
    
    try:
        await context.bot.send_message(
            chat_id=GROUP_CHAT_ID,
            text=group_message
        )
        logger.info(f"تم نشر التقرير #{report_id} في القروب")
    except Exception as e:
        logger.error(f"خطأ في إرسال التقرير للقروب: {e}")
        await query.message.reply_text("⚠️ تم حفظ التقرير لكن حدث خطأ في إرساله للقروب")
    
    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """إلغاء العملية الحالية"""
    context.user_data.clear()
    await update.message.reply_text(
        "تم إلغاء العملية",
        reply_markup=ReplyKeyboardMarkup([[]], resize_keyboard=True)
    )
    return ConversationHandler.END


async def summary_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """عرض قائمة الحصر"""
    user_id = update.effective_user.id
    
    if user_id not in ADMIN_IDS:
        await update.message.reply_text("⚠️ هذا الأمر متاح للمدير فقط")
        return
    
    keyboard = [
        [InlineKeyboardButton("📅 حصر اليوم + Excel", callback_data='summary_today')],
        [InlineKeyboardButton("📆 حصر الأسبوع + Excel", callback_data='summary_week')],
        [InlineKeyboardButton("📊 حصر الشهر + Excel", callback_data='summary_month')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "اختر نوع التقرير:",
        reply_markup=reply_markup
    )


async def generate_summary(update: Update, context: ContextTypes.DEFAULT_TYPE, period: str):
    """توليد ملف Excel وإرساله"""
    if update.message:
        chat_id = update.message.chat_id
        message = update.message
    else:
        query = update.callback_query
        await query.answer()
        chat_id = query.message.chat_id
        message = query.message
    
    await message.reply_text("⏳ جاري إنشاء التقرير...")
    
    try:
        reports = get_reports_by_period(period)
        
        if not reports:
            await message.reply_text(f"⚠️ لا توجد تقارير في هذه الفترة ({period})")
            return
        
        filename = create_excel_report(period, reports)
        
        period_names = {
            'today': 'اليوم',
            'week': 'الأسبوع',
            'month': 'الشهر'
        }
        
        caption = f"📊 تقرير {period_names[period]}\n📈 عدد التقارير: {len(reports)}"
        
        with open(filename, 'rb') as file:
            await context.bot.send_document(
                chat_id=GROUP_CHAT_ID,
                document=file,
                caption=caption,
                filename=filename
            )
        
        await message.reply_text(f"✅ تم إرسال تقرير {period_names[period]} إلى القروب")
        
        os.remove(filename)
        
    except Exception as e:
        logger.error(f"خطأ في إنشاء التقرير: {e}")
        await message.reply_text(f"❌ حدث خطأ في إنشاء التقرير:\n{str(e)}")


async def summary_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تقرير اليوم"""
    user_id = update.effective_user.id
    if user_id not in ADMIN_IDS:
        await update.message.reply_text("⚠️ هذا الأمر متاح للمدير فقط")
        return
    await generate_summary(update, context, 'today')


async def summary_week(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تقرير الأسبوع"""
    user_id = update.effective_user.id
    if user_id not in ADMIN_IDS:
        await update.message.reply_text("⚠️ هذا الأمر متاح للمدير فقط")
        return
    await generate_summary(update, context, 'week')


async def summary_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تقرير الشهر"""
    user_id = update.effective_user.id
    if user_id not in ADMIN_IDS:
        await update.message.reply_text("⚠️ هذا الأمر متاح للمدير فقط")
        return
    await generate_summary(update, context, 'month')


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """معالجة الضغط على الأزرار"""
    query = update.callback_query
    
    if query.data.startswith('summary_'):
        period = query.data.replace('summary_', '')
        await generate_summary(update, context, period)


def main():
    """البرنامج الرئيسي"""
    if not BOT_TOKEN:
        logger.error("خطأ: BOT_TOKEN غير موجود")
        print("❌ خطأ: BOT_TOKEN غير موجود في Environment Variables")
        return
    
    if not GROUP_CHAT_ID:
        logger.error("خطأ: GROUP_CHAT_ID غير موجود")
        print("❌ خطأ: GROUP_CHAT_ID غير موجود في Environment Variables")
        return
    
    init_database()
    
    application = Application.builder().token(BOT_TOKEN).build()
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('report', start_report)],
        states={
            SUPERVISOR_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_supervisor_name)],
            VISIT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_visit_date)],
            SCHOOL_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_school_name)],
            MAINTENANCE_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_maintenance_notes)],
            AC_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_ac_notes)],
            CLEANING_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_cleaning_notes)],
            REVIEW_REPORT: [CallbackQueryHandler(confirm_report)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    
    application.add_handler(CommandHandler('start', start))
    application.add_handler(conv_handler)
    application.add_handler(CommandHandler('summary', summary_menu))
    application.add_handler(CommandHandler('summary_today', summary_today))
    application.add_handler(CommandHandler('summary_week', summary_week))
    application.add_handler(CommandHandler('summary_month', summary_month))
    application.add_handler(CallbackQueryHandler(button_callback))
    
    logger.info("🤖 البوت يعمل الآن...")
    print("🤖 البوت يعمل الآن...")
    
    application.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)


if __name__ == '__main__':
    main()
